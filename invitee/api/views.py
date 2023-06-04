from invitee.models import Invitee
from rest_framework.views import APIView
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, CreateAPIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
import openpyxl
from rest_framework import status
from .serializers import BulkCreateSerializer, InviteeSerializer
from events.models import Event
import json

class InviteeListCreateListAPI(APIView):
    queryset = Invitee.objects.all()
    serializer_class = InviteeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['recognized', 'sent_email', "created_by"]
    search_fields = ["user__username", 'name', 'phone_number', 'unique_id', "name"]
    ordering_fields = ['name', 'name', "unique_id"]
    
    def get_queryset(self):
        queryset = Invitee.objects.all()

        # Apply the filters
        queryset = self.filter_queryset(queryset)

        # Apply ordering
        queryset = self.order_queryset(queryset)

        return queryset

    def filter_queryset(self, queryset):
        for backend in self.filter_backends:
            queryset = backend().filter_queryset(self.request, queryset, self)
        return queryset

    def order_queryset(self, queryset):
        ordering = self.request.query_params.get('ordering', None)
        if ordering:
            fields = [field.strip() for field in ordering.split(',')]
            return queryset.order_by(*fields)
        return queryset
    
    def get(self, request, event_pk):
        queryset = self.get_queryset().filter(event_id=event_pk)

        # Perform search if search query parameter is provided
        search_query = request.query_params.get('search', None)
        if search_query:
            search_backend = filters.SearchFilter()
            queryset = search_backend.filter_queryset(request, queryset, self)

        serializer = self.serializer_class(queryset, many=True)
        return Response(serializer.data)
    
    def post(self, request, event_pk):
        serializer = self.serializer_class(data=request.data)
        user = request.user
        event = Event.objects.prefetch_related("created_by").get(pk=event_pk)
        if serializer.is_valid() and event.created_by == user:
            serializer.save(created_by=request.user, event=event)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class InviteeRetriveUpdateAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Invitee.objects.all()
    serializer_class = InviteeSerializer
    permission_classes = [IsAuthenticated]
    lookup_field= "pk"
    
    
    def perform_update(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        
        
import random
import string
def random_string_generator(size=10, chars=string.ascii_lowercase + string.digits + string.ascii_uppercase):
    return ''.join(random.choice(chars) for _ in range(size))  
  
class InviteeBulkCreateAPI(CreateAPIView):
    queryset = Invitee.objects.all()
    serializer_class = BulkCreateSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        user=self.request.user
        if serializer.is_valid():
            try:
                event_instance = Event.objects.exclude(removed=True).get(pk=serializer.validated_data["event_id"], created_by=user)
                
                file = serializer.validated_data("file")
                print(file)
                
                wb = openpyxl.load_workbook(file)
                worksheet = wb[wb.sheetnames[0]]
                
                excel_data = list()
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value).lower())
                    excel_data.append(row_data)
                    
                features = excel_data.pop(0)
                
                name_index = features.index("name")
                email_index = features.index("email")

                try:
                    phonenumber_index = features.index("phonenumber")
                    features.pop(phonenumber_index)
                except:
                    phonenumber_index = 0
                features.pop(name_index)
                features.pop(email_index)
                non_important_features = []
                
                for i in features:
                    non_important_features.append(features.index(i))
                
                objs = []
                dict = {}
                for li in excel_data:
                    for extra in non_important_features:
                        dict[features[extra]] = li[extra]
                    extra = json.dumps(extra)
                    objs.append(Invitee(created_by=user, event=event_instance, name=li[name_index], email=li[email_index], phone_number=li[phonenumber_index] if phonenumber_index else 00000000, other_info=extra ,unique_id=random_string_generator(size=12)))
                
                # Invitee.objects.filter(event=event_instance).delete()
                Invitee.objects.bulk_create(objs)
                qs = Invitee.objects.filter(event=event_instance)
                event_instance.invitees.clear()
                event_instance.invitees.add(*qs)
                event_instance.save()
                return Response({"data":"added sucsessfully", "error":0, "code":2000})
            except Exception as e:
                return Response({"data":f"error {e}", "error":1, "code":4000})        


class delete_all_invitees(APIView):
    queryset = Invitee.objects.all()
    serializer_class = BulkCreateSerializer
    permission_classes = [IsAuthenticated]
    def post(self, request, pk):
        eo = Event.objects.get(pk=pk)
        if eo.created_by == request.user:
            for i in eo.invitees.all():
                i.delete()
            return Response({"data":"deleted sucsessfully", "error":0, "code":2000})
        else:
            return Response({"data":"not same user", "error":1, "code":4000})