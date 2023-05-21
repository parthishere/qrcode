from rest_framework.response import Response
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, CreateAPIView
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from ..models import Event
from accounts.models import Invitee
from .serializers import EventSerializer, BulkCreateSerializer, InviteeSerializer
from rest_framework.permissions import IsAuthenticated
import openpyxl
import json
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.contrib.auth.decorators import login_required

class EventListCreateAPI(ListCreateAPIView):
    queryset = Event.objects.all()
    serializer_class = EventSerializer
    permission_classes = [IsAuthenticated,]
    lookup_field = ['pk']
    lookup_url_kwarg = ['pk']
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def list(self, request):
        queryset = self.get_queryset().filter(created_by=request.user).exclude(removed=True)
        serializer = EventSerializer(queryset, many=True)
        return Response(serializer.data)
    

class EventRetriveUpdateAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Event.objects.all()
    serializer_class = EventSerializer
    permission_classes = [IsAuthenticated]
    lookup_field= "pk"
    
    
    def perform_update(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        
    def perform_destroy(self, instance):
        if self.request.user == instance.created_by:
            instance.delete()
        
        
            
class InviteeListCreateListAPI(APIView):
    queryset = Invitee.objects.all()
    serializer_class = InviteeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['recognized', 'sent_email', "created_by"]
    search_fields = ["user__username", 'name', 'phone_number', 'unique_id', "email"]
    
    def get(self, request, event_pk):
        user = request.user
        queryset = self.get_queryset().filter(event__invitee=user)
        serializer = self.serializer_class(queryset, many=True)
        return Response(serializer.data)
    
class InviteeRetriveUpdateAPIView(RetrieveUpdateDestroyAPIView):
    queryset = Invitee.objects.all()
    serializer_class = InviteeSerializer
    permission_classes = [IsAuthenticated]
    lookup_field= "pk"
    
    
    def perform_update(self, serializer):
        instance = serializer.save(created_by=self.request.user)
    
  
import random
import string
def random_string_generator(size=10, chars=string.ascii_lowercase + string.digits + string.ascii_uppercase):
    return ''.join(random.choice(chars) for _ in range(size))  
  
class InviteeBulkCreateAPI(CreateAPIView):
    queryset = Invitee.objects.all()
    serializer_class = BulkCreateSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        user=self.request.user
        if serializer.is_valid():
            try:
                event_instance = Event.objects.exclude(removed=True).get(pk=serializer.validated_data["event_id"], created_by=user)
                
                file = serializer.validated_data("file")
                print(file)
                
                wb = openpyxl.load_workbook(file)
                worksheet = wb[wb.sheetnames[0]]
                
                excel_data = list()
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value).lower())
                    excel_data.append(row_data)
                    
                features = excel_data.pop(0)
            
                name_index = features.index("name")
                email_index = features.index("email")

                try:
                    phonenumber_index = features.index("phonenumber")
                    features.pop(phonenumber_index)
                except:
                    phonenumber_index = 0
                features.pop(name_index)
                features.pop(email_index)
                non_important_features = []
                
                for i in features:
                    non_important_features.append(features.index(i))
                
                objs = []
                dict = {}
                for li in excel_data:
                    for extra in non_important_features:
                        dict[features[extra]] = li[extra]
                    extra = json.dumps(extra)
                    objs.append(Invitee(created_by=user, event=event_instance, name=li[name_index], email=li[email_index], phone_number=li[phonenumber_index] if phonenumber_index else 00000000, other_info=extra ,unique_id=random_string_generator(size=12)))
                
                # Invitee.objects.filter(event=event_instance).delete()
                Invitee.objects.bulk_create(objs)
                qs = Invitee.objects.filter(event=event_instance)
                # event_instance.invitees.clear()
                event_instance.invitees.add(*qs)
                event_instance.save()
                return Response({"data":"added sucsessfully", "error":0, "code":2000})
            except Exception as e:
                return Response({"data":f"error {e}", "error":1, "code":4000})        
         

class delete_all_invitees(APIView):
    queryset = Invitee.objects.all()
    serializer_class = BulkCreateSerializer
    permission_classes = [IsAuthenticated]
    def post(self, request, pk):
        eo = Event.objects.get(pk=pk)
        if eo.created_by == request.user:
            for i in eo.invitees.all():
                i.delete()
            return Response({"data":"deleted sucsessfully", "error":0, "code":2000})
        else:
            return Response({"data":"not same user", "error":1, "code":4000})
            
  
@api_view(["POST"]) 
@login_required         
def scan(request, event_pk=None):
    """
    
    Send Intrument found_string in the request.body as a json response 
    @param: "found_string"
    
    const data = {
        "found_string":"qr_name, qr_email, qr_phone_number, qr_event_primary_key, qr_unique_id"
    }
    
    fetch("URI", {
        method:"POST",
        headers: {
            "Content-Type":"application/json"
        },
        body: JSON.stringyfy(data)
    })
    
    .then(response => response.json())
    
    
    """
    
    try:
        data = json.loads(request.body)
        string = data["found_string"]
        qr_name, qr_email, qr_phone_number, qr_event_primary_key, qr_unique_id = string.split(',')[0], string.split(',')[1], string.split(',')[2], string.split(',')[3], string.split(',')[4]
        
        
    except Exception as e:
        print(e)
        print("will not recognize")
        data = "Not valid QR"
        message = {"message": "not valid QR", "code": 1004}
        return Response(message)
    
    try:
        og_event = Event.objects.prefetch_related("recognized_invitees").get(pk=event_pk, created_by=request.user)
        
        q = Invitee.objects.get(email=qr_email, unique_id=qr_unique_id, event=og_event)

        q_in_event_exists = True if q.event == og_event else False 

        q_already_scaned = q.recognized
    except:
        message = {"message": "not valid event", "code": 1004}
        return Response(message)
    
    
    try:
        if q_in_event_exists and int(event_pk) == int(qr_event_primary_key):                
            
        
            if q_already_scaned:
                message = {"message": "Scanned Again", "code": 1000}
                return Response(message)
            else:
               
                q.recognized = True
                q.save()
                og_event.recognized_invitees.add(q)
                og_event.save()
                
                message = {"message": {"qrcode":qr_name, "qr_email":qr_email, "qr_phone_number":qr_phone_number, "event_pk":og_event.pk, "qr_unique_id":qr_unique_id}, "code": "1001" }
                return Response(message)
        else:
            data = "User does not exists in the event"
            message = {"message": "User does not exists in the event", "code": 1002}
            return Response(message)
            
    except Exception as e:
        message = {"message": f"Something Is Wrong either Event or Invitee doesn't exist {e}", "code": 1003}
    return Response(message)

from django.template.loader import render_to_string
from django.conf import settings

from django.core.mail import EmailMessage

from mimetypes import guess_type
from os.path import basename


@api_view(["GET"])
def send_email_to_all(request, event_pk):
    print("heya in the task")
    user = request.user
    event_instance = Event.objects.exclude(removed=True).prefetch_related("invitees").get(pk=event_pk, created_by=user)
    
    # connection = mail.get_connection()
    # connection.open()
    invitees = [i for i in event_instance.invitees.all()]        
    all_mail = []  
    for invitee in invitees :
        invitee.save()
        if invitee.qr_code:
            template = render_to_string('app/email_template.html', {"extra_fields":''.join(random.choice(string.ascii_lowercase + string.digits + string.ascii_uppercase) for _ in range(10)), 'name': invitee.name, 'email':invitee.email, 'phone_number':invitee.phone_number, "event_name":event_instance.event_name, "event_date":event_instance.event_date, "about":event_instance.about, "created_by":event_instance.created_by, "organization_or_college":event_instance.organization_or_college, "unique_id": invitee.unique_id})
            invitee.sent_email = True
            invitee.save()
            subject = "Initation for: " + event_instance.event_name
            message = template
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [invitee.email]
       
            image = invitee.qr_code
    
            
            image.open()

            # email = EmailMessage("sub", "pub", "thakkar.parth782@gmail.com", ["parthishere1234@gmail.com"])
            email = EmailMessage(subject, message, email_from, recipient_list,)
            email.attach(basename(image.name),image.read(), guess_type(image.name)[0])
            email.send(fail_silently=False)
            all_mail.append(email)
            image.close()
            
       
  
    # connection.send_messages(all_mail)
    # connection.close()
    message = {"message": "done", "code": 200}
    return Response(message)
        
    
@api_view(["GET"]) 
@login_required
def send_email_to_remaining(request, event_pk):
    
    user = request.user
    event_instance = Event.objects.exclude(removed=True).prefetch_related("invitees").get(pk=event_pk, created_by=user)
    
    invitees = event_instance.invitees.all()
    # print(participant)
    
    for invitee in invitees :
        if invitee.qr_code and not invitee.sent_email:
            template = render_to_string('app/email_template.html', {'name': invitee.name, 'email':invitee.email, 'phone_number':invitee.phone_number, "event_name":event_instance.event_name, "event_date":event_instance.event_date, "about":event_instance.about, "created_by":event_instance.created_by, "organization_or_college":event_instance.organization_or_college})
            invitee.sent_email = True
            invitee.save()
            subject = "Initation for: " + event_instance.event_name
            message = template
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [invitee.email,]
            mail = EmailMessage(subject, message, email_from, recipient_list)
            mail.attach(invitee.qr_code.name, invitee.qr_code.read())
            mail.send()
       
    # group_name = get_user_name()  # Find out way to get same as what is printed on connect()

    # channel_layer = get_channel_layer()
    # # Trigger reload message sent to group
    # async_to_sync(channel_layer.group_send)(
    #     group_name,
    #     {'type': 'reload_page'}
    # )
    message = {"message": "done", "code": 200}
    return Response(message)