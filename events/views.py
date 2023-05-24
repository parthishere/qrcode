from enum import unique
from django.shortcuts import render, redirect
from django.urls import reverse
from events.forms import EventForm, FileForm, InviteeForm

from events.models import Event
from invitee.models import Invitee
from django.contrib.auth.decorators import login_required
from django.conf import settings
import openpyxl
import json

# Create your views here.
@login_required
def list_event(request):
    context = {}
    user = request.user
    events = Event.objects.filter(created_by=user).exclude(removed=True)
    context['events_list'] = events
    return render(request, "events/list-event.html", context)
   
   
@login_required 
def create_event(request):
    context = {}
    
    user = request.user
    events = Event.objects.filter(created_by=user).exclude(removed=True)
    event_count = events.count()
    
    form = EventForm(request.POST or None)
    
    context['form'] = form
    
    if request.POST:
        if event_count <= settings.MAX_FREE_EVENTS or user.is_pro:
            event_instance = form.save(commit=False)
            event_instance.created_by = user
            event_instance.save()
            context['form'] = form
            return redirect(event_instance.get_absolute_url)
        else:
            return redirect("event:become-pro")
        
    return render(request, "events/create-event.html", context)

@login_required 
def detail_event(request, pk=None):
    context = {}
    user = request.user
    event_instance = Event.objects.select_related("created_by").prefetch_related("moderators", "invitees").exclude(removed=True).get(pk=pk, removed=False)
    if user == event_instance.created_by:
        context['event_detail'] = event_instance
    else:
        return redirect("home")
    return render(request, "events/detail-event.html", context)

@login_required 
def update_event(request, pk=None):
    context = {}
    user = request.user
    event_instance = Event.objects.exclude(removed=True).get(pk=pk)
    
    form = EventForm(request.POST or None, instance=event_instance)
    
    context['form'] = form
    
    if request.POST:
        if user is event_instance.created_by and (event_instance.event_update_count <= settings.MAX_FREE_UPDATE_EVENTS or user.is_pro):
            event_instance = form.save(commit=False)
            event_instance.created_by = user
            
            if not user.is_pro:
                event_instance.event_update_count += 1
        
            event_instance.save()
            context['form'] = form
            return redirect(event_instance.get_absolute_url)
        else:
            return redirect("event:become-pro")
    return render(request, "events/create-event.html", context)
        
   
@login_required  
def delete_event(request, pk=None):
    user = request.user
    event_instance = Event.objects.exclude(removed=True).get(pk=pk)
  
    if user == event_instance.created_by:
        event_instance.removed = True
        event_instance.save()
        return redirect("events:list")
    
    return redirect(event_instance.get_absolute_url)
    
    
@login_required
def delete_event_pass(request, pk=None):
    user=request.user
    event_instance = Event.objects.exclude(removed=True).get(pk=pk, created_by=user)
    if event_instance.exists() and request.POST and (event_instance.event_update_count <= settings.MAX_FREE_UPDATE_EVENTS or user.is_pro):
        if not user.is_pro:
            event_instance.event_update_count += 1
            event_instance.save()
        event_instance.predefined_pass_number.delete() if event_instance.pre_define_pass else None
       
        return redirect(event_instance.get_absolute_url)
        
        
@login_required
def add_participant(request, pk=None):
    context = {}
    user=request.user
    event_instance = Event.objects.exclude(removed=True).get(pk=pk, created_by=user)
    form = InviteeForm(request.POST or None)
    context["form"] = form
    context['pk'] = event_instance.pk
    print(event_instance.created_by is user)
    if event_instance.created_by == user and request.POST and form.is_valid():
        instance = form.save(commit=False)
        instance.event = event_instance
        instance.created_by = user
        instance.save()
        if instance.recognized == True:
            event_instance.recognized_invitees.add(instance)
        event_instance.invitees.add(instance)
        event_instance.save()
        context['form'] = form
        return redirect(event_instance.get_absolute_url)
    
    return render(request, "events/create-participant.html", context)
        

import random
import string
def random_string_generator(size=10, chars=string.ascii_lowercase + string.digits + string.ascii_uppercase):
    return ''.join(random.choice(chars) for _ in range(size))

  
  
import pandas as pd     
@login_required 
def bulk_create(request, pk=None):
    context = {}
    user=request.user
    event_instance = Event.objects.exclude(removed=True).get(pk=pk, created_by=user)
    form = FileForm(request.POST or None, request.FILES or None)
    context['form'] = form
    context['idpk'] = event_instance.pk
    
    if request.POST and form.is_valid(): #and (event_instance.event_update_count <= settings.MAX_FREE_UPDATE_EVENTS or user.is_pro):

        file = request.FILES.get("excel_file")
        print(file)
        
        wb = openpyxl.load_workbook(file)
        worksheet = wb[wb.sheetnames[0]]
        
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value).lower())
            excel_data.append(row_data)
            
        features = excel_data.pop(0)
    
        name_index = features.index("name")
        email_index = features.index("email")

        # try:
        #     phonenumber_index = features.index("phonenumber")
        #     features.pop(phonenumber_index)
        # except:
        #     phonenumber_index = 0
        features.pop(name_index)
        features.pop(email_index)
        non_important_features = []
        
        for i in features:
            non_important_features.append(features.index(i))
        
        objs = []
        dict = {}
        for li in excel_data:
            for extra in non_important_features:
                dict[features[extra]] = li[extra]
            extra = json.dumps(extra)
            if li[name_index] and li[email_index] and li[name_index] != "none":
                print(li[name_index])
                objs.append(Invitee(created_by=user, event=event_instance, name=li[name_index], email=li[email_index], other_info=extra ,phone_number=0, unique_id=random_string_generator(size=12)))
        
        event_instance.invitees.clear()
        Invitee.objects.filter(event=event_instance).delete()
        Invitee.objects.bulk_create(objs)
        qs = Invitee.objects.filter(event=event_instance)
        
        event_instance.invitees.add(*qs)
        
        
        return redirect(event_instance.get_absolute_url)
    return render(request, "events/bulk-create.html", context)


def delete_participant(request, unique_id=None, pk=None):
    event_instance = Event.objects.exclude(removed=True).get(pk=pk)
    invitee = Invitee.objects.get(unique_id=unique_id)
    
    if invitee in event_instance.invitees.all():
        event_instance.invitees.remove(invitee)
    try:
        invitee.delete()
    except:
        pass
    event_instance.save()
    return redirect(event_instance.get_absolute_url)


def remove_all_participant(request, pk=None):
    event_instance = Event.objects.exclude(removed=True).get(pk=pk)
    event_instance.invitees.clear()
    return redirect(event_instance.get_absolute_url)
    
    
def see_all_recognized_participant(request, pk=None):
    context = {}
    event_instance = Event.objects.select_related("created_by").prefetch_related("moderators", "invitees").exclude(removed=True).get(pk=pk, removed=False)
    context['event_detail'] = event_instance
    
    return render(request, "events/recognized.html", context)

def see_all_unrecognized_participant(request, pk=None):
    context = {}
    event_instance = Event.objects.select_related("created_by").prefetch_related("moderators", "invitees").exclude(removed=True).get(pk=pk, removed=False)
    
    context['event_detail'] = event_instance
    
    return render(request, "events/unrecognized.html", context)